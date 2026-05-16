"""
Tally Excel Import — parsing and row classification.

The client exports inventory from Tally as an .xlsx with NO header row.
Data starts around row 3. Fixed columns:

    C = Brand          (e.g. "AMIRI", "Chanel", "AXEL ARIGATO^")
    D = Product name   (e.g. "A- 22 Hoodie White B176")
    E = Tally code     (e.g. "000B176")  -> unique dedup key
    F = Size           (e.g. "Large", "EU 41", "766467", blank)
    G = Category       (e.g. "SHOES", "T-SHIRT", "HOODIE")
    H = MRP            (e.g. "54,000.00/Nos")
    I = Quantity       (e.g. 1)
    J = Selling price  (e.g. "35,000.00/Nos")
    K = Selling price  (numeric twin of J, e.g. 35000)

This module only parses and classifies. Persistence lives in services.py.
"""

from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional

# 1-based column indices (openpyxl convention)
COL_BRAND = 3       # C
COL_NAME = 4        # D
COL_TALLY_CODE = 5  # E
COL_SIZE = 6        # F
COL_CATEGORY = 7    # G
COL_MRP = 8         # H
COL_QUANTITY = 9    # I
COL_SELLING_STR = 10  # J
COL_SELLING_NUM = 11  # K

STATUS_NEW = "new"
STATUS_DUPLICATE = "duplicate"
STATUS_ERROR = "error"


def _clean_text(value: Any) -> str:
    """Trim whitespace and strip trailing junk like the '^' some brands carry."""
    if value is None:
        return ""
    text = str(value).strip()
    # Drop stray non-alphanumeric trailing characters (e.g. "AXEL ARIGATO^")
    text = re.sub(r"[^\w\)\.\&\-\s]+$", "", text).strip()
    return text


def parse_price(value: Any) -> Optional[Decimal]:
    """
    Parse a Tally price cell into a Decimal.

    Accepts numeric cells (35000) and strings like "1,10,000.00/Nos".
    Returns None if it cannot be parsed.
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    text = str(value).strip()
    # Strip unit suffix like "/Nos", "/Pcs"
    text = re.sub(r"/\s*[A-Za-z]+\s*$", "", text).strip()
    # Remove thousands separators (Indian or western grouping) — only commas
    text = text.replace(",", "")
    # Keep digits and a single decimal point
    text = re.sub(r"[^\d.]", "", text)
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def parse_quantity(value: Any) -> int:
    """Parse the quantity cell. Blank/garbage -> 0."""
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        try:
            return max(0, int(value))
        except (ValueError, TypeError):
            return 0
    text = re.sub(r"[^\d-]", "", str(value))
    if not text or text == "-":
        return 0
    try:
        return max(0, int(text))
    except (ValueError, TypeError):
        return 0


def parse_tally_workbook(file_obj) -> List[Dict[str, Any]]:
    """
    Read the uploaded .xlsx and return a list of normalized row dicts.

    Each dict: row, brand, name, tally_code, size, category,
               mrp, selling_price, quantity, error (optional str)

    Rows that are completely blank are skipped silently. Rows that have
    content but fail validation get an 'error' message and are still
    returned so the UI can show them.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: List[Dict[str, Any]] = []
    for idx, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        # raw is a tuple; column N is raw[N-1]
        def cell(col_1based: int) -> Any:
            i = col_1based - 1
            return raw[i] if i < len(raw) else None

        brand = _clean_text(cell(COL_BRAND))
        name = _clean_text(cell(COL_NAME))
        tally_code = _clean_text(cell(COL_TALLY_CODE))
        size = _clean_text(cell(COL_SIZE))
        category = _clean_text(cell(COL_CATEGORY))

        # Skip fully empty rows (header gap / trailing blanks)
        if not any([brand, name, tally_code, category]):
            continue

        mrp = parse_price(cell(COL_MRP))
        selling = parse_price(cell(COL_SELLING_NUM))
        if selling is None:
            selling = parse_price(cell(COL_SELLING_STR))
        quantity = parse_quantity(cell(COL_QUANTITY))

        error: Optional[str] = None
        if not name:
            error = "Missing product name"
        elif not tally_code:
            error = "Missing Tally code"
        elif selling is None and mrp is None:
            error = "No valid price (MRP or selling price)"

        # Sensible fallbacks so a row with one price still imports
        if selling is None:
            selling = mrp
        if mrp is None:
            mrp = selling

        rows.append(
            {
                "row": idx,
                "brand": brand or "Unknown",
                "name": name,
                "tally_code": tally_code,
                "size": size or None,
                "category": category or "Uncategorized",
                "mrp": str(mrp) if mrp is not None else None,
                "selling_price": str(selling) if selling is not None else None,
                "quantity": quantity,
                "error": error,
            }
        )

    return rows


def classify_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Tag each parsed row with a status:
      - error     : failed validation
      - duplicate : a non-deleted Product already has this Tally code
      - new       : will be created on commit

    Duplicate detection is by Product.product_code == tally_code, which
    makes re-uploading the same sheet idempotent (existing rows skipped).
    """
    from .models import Product

    codes = [r["tally_code"] for r in rows if r["tally_code"] and not r["error"]]
    existing = set(
        Product.objects.filter(
            product_code__in=codes, is_deleted=False
        ).values_list("product_code", flat=True)
    )

    seen_in_file: set = set()
    classified: List[Dict[str, Any]] = []
    for r in rows:
        item = dict(r)
        if r["error"]:
            item["status"] = STATUS_ERROR
            item["message"] = r["error"]
        elif r["tally_code"] in existing or r["tally_code"] in seen_in_file:
            item["status"] = STATUS_DUPLICATE
            item["message"] = "Already exists — skipped"
        else:
            item["status"] = STATUS_NEW
            item["message"] = ""
            seen_in_file.add(r["tally_code"])
        classified.append(item)
    return classified


def summarize(classified: List[Dict[str, Any]]) -> Dict[str, int]:
    summary = {
        "total": len(classified),
        "new": 0,
        "duplicate": 0,
        "error": 0,
    }
    for r in classified:
        summary[r["status"]] = summary.get(r["status"], 0) + 1
    return summary
